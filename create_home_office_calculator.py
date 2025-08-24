from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_home_office_calculator():
    wb = Workbook()
    
    # Create Simplified Method sheet
    ws_simplified = wb.active
    ws_simplified.title = "Simplified Method"
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    input_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    result_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Simplified Method Calculator
    simplified_data = [
        ["HOME OFFICE DEDUCTION - SIMPLIFIED METHOD", "", ""],
        ["", "", ""],
        ["STEP 1: MEASURE YOUR HOME OFFICE", "", ""],
        ["Home office square footage:", "", ""],
        ["(Maximum 300 sq ft for simplified method)", "", ""],
        ["", "", ""],
        ["STEP 2: CALCULATION", "", ""],
        ["Office square footage", "=B4", ""],
        ["Rate per square foot", "$5.00", "(IRS Standard Rate)"],
        ["Total deduction", "=B8*5", ""],
        ["", "", ""],
        ["STEP 3: ANNUAL DEDUCTION LIMIT", "", ""],
        ["Your home office deduction", "=B10", ""],
        ["Maximum allowed (300 sq ft)", "$1,500", ""],
        ["Your eligible deduction", "=MIN(B12,B13)", ""],
        ["", "", ""],
        ["ADVANTAGES OF SIMPLIFIED METHOD:", "", ""],
        ["• No need to track home expenses", "", ""],
        ["• No depreciation to recapture when selling", "", ""],
        ["• Simple calculation", "", ""],
        ["• Less record keeping required", "", ""],
        ["", "", ""],
        ["REQUIREMENTS:", "", ""],
        ["• Use home office regularly and exclusively for business", "", ""],
        ["• Office must be principal place of business OR", "", ""],
        ["• Used regularly to meet clients/customers", "", ""],
    ]
    
    # Add simplified method data
    for row_idx, row_data in enumerate(simplified_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_simplified.cell(row=row_idx, column=col_idx)
            if isinstance(value, str) and value.startswith("="):
                cell.value = value
            else:
                cell.value = value
            
            # Apply formatting
            if row_idx == 1:  # Title
                cell.font = Font(bold=True, size=16, color="2E86AB")
            elif row_idx in [3, 7, 12, 17, 23]:  # Section headers
                cell.font = Font(bold=True, color="2E86AB")
            elif row_idx == 4:  # Input cell
                cell.fill = input_fill
                cell.border = border
                if col_idx == 2:
                    cell.value = 150  # Sample value
            elif row_idx in [10, 14]:  # Result cells
                cell.fill = result_fill
                cell.border = border
                cell.font = Font(bold=True)
    
    # Set column widths
    ws_simplified.column_dimensions['A'].width = 35
    ws_simplified.column_dimensions['B'].width = 15
    ws_simplified.column_dimensions['C'].width = 25
    
    # Create Actual Expense Method sheet
    ws_actual = wb.create_sheet("Actual Expense Method")
    
    actual_data = [
        ["HOME OFFICE DEDUCTION - ACTUAL EXPENSE METHOD", "", ""],
        ["", "", ""],
        ["STEP 1: HOME MEASUREMENTS", "", ""],
        ["Total home square footage:", "", ""],
        ["Home office square footage:", "", ""],
        ["Business use percentage:", "=B5/B4*100", "%"],
        ["", "", ""],
        ["STEP 2: ANNUAL HOME EXPENSES", "Amount", "Business Portion"],
        ["Mortgage interest", "", "=B9*B6/100"],
        ["Property taxes", "", "=B10*B6/100"],
        ["Utilities (electric, gas, water)", "", "=B11*B6/100"],
        ["Home insurance", "", "=B12*B6/100"],
        ["Repairs & maintenance", "", "=B13*B6/100"],
        ["Depreciation (if owned)", "", "=B14*B6/100"],
        ["HOA fees", "", "=B15*B6/100"],
        ["Security system", "", "=B16*B6/100"],
        ["", "", ""],
        ["STEP 3: DIRECT OFFICE EXPENSES", "", "Business Portion"],
        ["Office supplies", "", "=B19"],
        ["Office equipment", "", "=B20"],
        ["Office furniture", "", "=B21"],
        ["Office repairs", "", "=B22"],
        ["", "", ""],
        ["STEP 4: TOTAL DEDUCTION", "", ""],
        ["Total indirect expenses", "=SUM(C9:C16)", ""],
        ["Total direct expenses", "=SUM(C19:C22)", ""],
        ["Total home office deduction", "=B25+B26", ""],
        ["", "", ""],
        ["COMPARISON WITH SIMPLIFIED METHOD", "", ""],
        ["Actual expense method", "=B27", ""],
        ["Simplified method (from other sheet)", "='Simplified Method'.B14", ""],
        ["Recommended method", "=IF(B29>B30,\"Actual Expense\",\"Simplified\")", ""],
        ["Additional savings vs simplified", "=MAX(0,B29-B30)", ""],
    ]
    
    # Add actual expense data
    for row_idx, row_data in enumerate(actual_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_actual.cell(row=row_idx, column=col_idx)
            if isinstance(value, str) and value.startswith("="):
                cell.value = value
            else:
                cell.value = value
            
            # Apply formatting
            if row_idx == 1:  # Title
                cell.font = Font(bold=True, size=16, color="2E86AB")
            elif row_idx in [3, 8, 18, 24, 28]:  # Section headers
                cell.font = Font(bold=True, color="2E86AB")
            elif row_idx in [4, 5] and col_idx == 2:  # Input cells
                cell.fill = input_fill
                cell.border = border
                if row_idx == 4:
                    cell.value = 2000  # Sample total home sq ft
                elif row_idx == 5:
                    cell.value = 200   # Sample office sq ft
            elif row_idx in range(9, 23) and col_idx == 2:  # Expense input cells
                cell.fill = input_fill
                cell.border = border
            elif row_idx in [6, 27, 29, 30, 31, 32]:  # Result cells
                cell.fill = result_fill
                cell.border = border
                cell.font = Font(bold=True)
    
    # Add sample data for actual expenses
    sample_expenses = {
        9: 12000,   # Mortgage interest
        10: 3000,   # Property taxes
        11: 2400,   # Utilities
        12: 1200,   # Insurance
        13: 800,    # Repairs
        14: 2000,   # Depreciation
        19: 300,    # Office supplies
        20: 1500,   # Office equipment
    }
    
    for row, amount in sample_expenses.items():
        ws_actual.cell(row=row, column=2).value = amount
    
    # Set column widths
    ws_actual.column_dimensions['A'].width = 35
    ws_actual.column_dimensions['B'].width = 15
    ws_actual.column_dimensions['C'].width = 20
    
    # Create Documentation Checklist sheet
    ws_docs = wb.create_sheet("Documentation Checklist")
    
    doc_data = [
        ["HOME OFFICE DOCUMENTATION CHECKLIST", "", ""],
        ["", "", ""],
        ["✓", "REQUIRED DOCUMENTATION", "NOTES"],
        ["", "", ""],
        ["□", "Photos of home office space", "Show exclusive business use"],
        ["□", "Floor plan or measurements", "Prove square footage calculations"],
        ["□", "Utility bills for full year", "Electric, gas, water, internet"],
        ["□", "Property tax statements", "Annual property tax bill"],
        ["□", "Mortgage interest statements", "Form 1098 from lender"],
        ["□", "Home insurance declarations", "Annual insurance coverage"],
        ["□", "Receipts for office supplies", "Business-only purchases"],
        ["□", "Receipts for office equipment", "Computers, printers, furniture"],
        ["□", "Repair and maintenance receipts", "Home office specific repairs"],
        ["□", "HOA fee statements", "If applicable"],
        ["□", "Security system bills", "If used for business"],
        ["□", "Depreciation calculations", "If claiming depreciation"],
        ["", "", ""],
        ["EXCLUSIVE USE TEST:", "", ""],
        ["□", "Office used ONLY for business", "No personal use allowed"],
        ["□", "Regular and continuous use", "Not occasional or incidental"],
        ["□", "Principal place of business OR", "Main business location"],
        ["□", "Regular client/customer meetings", "If not principal location"],
        ["", "", ""],
        ["RECORD KEEPING TIPS:", "", ""],
        ["• Keep records for at least 3 years", "", ""],
        ["• Store digital copies in cloud", "", ""],
        ["• Organize by tax year", "", ""],
        ["• Take photos of office setup annually", "", ""],
        ["• Maintain business calendar/log", "", ""],
        ["", "", ""],
        ["⚠️ COMMON MISTAKES TO AVOID:", "", ""],
        ["• Using office for personal activities", "", ""],
        ["• Claiming entire home as office", "", ""],
        ["• Not maintaining proper records", "", ""],
        ["• Mixing personal and business expenses", "", ""],
        ["• Forgetting to track depreciation recapture", "", ""],
    ]
    
    for row_idx, row_data in enumerate(doc_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_docs.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            if row_idx == 1:  # Title
                cell.font = Font(bold=True, size=16, color="2E86AB")
            elif row_idx in [3, 18, 24, 31]:  # Section headers
                cell.font = Font(bold=True, color="2E86AB")
            elif col_idx == 1 and value == "□":  # Checkboxes
                cell.font = Font(size=14)
                cell.alignment = Alignment(horizontal='center')
            elif value and (value.startswith("•") or value.startswith("⚠️")):
                cell.font = Font(bold=True)
    
    # Set column widths
    ws_docs.column_dimensions['A'].width = 5
    ws_docs.column_dimensions['B'].width = 40
    ws_docs.column_dimensions['C'].width = 30
    
    # Create Instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ["TAXFIX HOME OFFICE DEDUCTION CALCULATOR", "", ""],
        ["", "", ""],
        ["How to Choose the Right Method:", "", ""],
        ["", "", ""],
        ["1. SIMPLIFIED METHOD:", "", ""],
        ["   • Best for: Small offices, renters, simple situations", "", ""],
        ["   • Maximum: 300 sq ft ($1,500 deduction)", "", ""],
        ["   • Rate: $5 per square foot", "", ""],
        ["   • Pros: Easy, no depreciation recapture", "", ""],
        ["   • Cons: Lower deduction for large offices", "", ""],
        ["", "", ""],
        ["2. ACTUAL EXPENSE METHOD:", "", ""],
        ["   • Best for: Large offices, homeowners, high expenses", "", ""],
        ["   • Calculation: Business % × Total home expenses", "", ""],
        ["   • Pros: Potentially higher deduction", "", ""],
        ["   • Cons: More record keeping, depreciation recapture", "", ""],
        ["", "", ""],
        ["3. HOW TO USE THIS CALCULATOR:", "", ""],
        ["   • Fill in blue cells with your information", "", ""],
        ["   • Review both methods on their respective sheets", "", ""],
        ["   • Compare results to choose best method", "", ""],
        ["   • Use Documentation Checklist for records", "", ""],
        ["", "", ""],
        ["4. BUSINESS USE PERCENTAGE:", "", ""],
        ["   • Office sq ft ÷ Total home sq ft", "", ""],
        ["   • Must be used exclusively for business", "", ""],
        ["   • Regular and continuous use required", "", ""],
        ["", "", ""],
        ["5. QUALIFYING REQUIREMENTS:", "", ""],
        ["   • Exclusive business use (no personal use)", "", ""],
        ["   • Principal place of business OR", "", ""],
        ["   • Regular meetings with clients/customers", "", ""],
        ["   • Administrative/management activities", "", ""],
        ["", "", ""],
        ["⚠️ IMPORTANT WARNINGS:", "", ""],
        ["• Depreciation recapture when selling home (actual method)", "", ""],
        ["• Must pass exclusive use test", "", ""],
        ["• Keep detailed records for audit protection", "", ""],
        ["• Consider professional tax advice for complex situations", "", ""],
        ["• Review IRS Publication 587 for complete rules", "", ""],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            if row_idx == 1:  # Title
                cell.font = Font(bold=True, size=16, color="2E86AB")
            elif ":" in str(value) and not value.startswith(" "):
                cell.font = Font(bold=True, color="2E86AB")
            elif value and (value.startswith("•") or value.startswith("⚠️")):
                cell.font = Font(bold=True)
    
    # Set column widths
    ws_instructions.column_dimensions['A'].width = 50
    ws_instructions.column_dimensions['B'].width = 20
    ws_instructions.column_dimensions['C'].width = 20
    
    # Save the workbook
    wb.save("C:/Users/Ben/taxfix-updated-site/TaxFix_Home_Office_Deduction_Calculator_2025.xlsx")
    print("Home Office Deduction Calculator created successfully!")

if __name__ == "__main__":
    create_home_office_calculator()