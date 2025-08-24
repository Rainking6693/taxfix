import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar

def create_mileage_log():
    wb = Workbook()
    
    # Create main tracking sheet
    ws_main = wb.active
    ws_main.title = "Mileage Log"
    
    # Headers with styling
    headers = [
        "Date", "Start Location", "End Location", "Trip Purpose", 
        "Start Odometer", "End Odometer", "Total Miles", 
        "Business %", "Business Miles", "Personal Miles"
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    column_widths = [12, 25, 25, 30, 15, 15, 12, 12, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws_main.column_dimensions[get_column_letter(i)].width = width
    
    # Add sample data and formulas for first few rows
    sample_data = [
        ["1/1/2025", "Home", "Client Office", "Business Meeting", 12000, 12025, "", 100, "", ""],
        ["1/2/2025", "Home", "Grocery Store", "Personal Shopping", 12025, 12035, "", 0, "", ""],
        ["1/3/2025", "Home", "Conference Center", "Business Conference", 12035, 12085, "", 100, "", ""]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_main.cell(row=row_idx, column=col_idx).value = value
            ws_main.cell(row=row_idx, column=col_idx).border = border
    
    # Add formulas for calculations (rows 2-51 for a full year)
    for row in range(2, 52):
        # Total Miles formula (End Odometer - Start Odometer)
        ws_main[f'G{row}'] = f'=IF(AND(E{row}<>"",F{row}<>""),F{row}-E{row},"")'
        
        # Business Miles formula (Total Miles * Business %)
        ws_main[f'I{row}'] = f'=IF(AND(G{row}<>"",H{row}<>""),G{row}*H{row}/100,"")'
        
        # Personal Miles formula (Total Miles - Business Miles)
        ws_main[f'J{row}'] = f'=IF(AND(G{row}<>"",I{row}<>""),G{row}-I{row},"")'
        
        # Add borders
        for col in range(1, 11):
            ws_main.cell(row=row, column=col).border = border
    
    # Monthly Summary Section
    ws_summary = wb.create_sheet("Monthly Summary")
    
    # Summary headers
    summary_headers = ["Month", "Total Miles", "Business Miles", "Personal Miles", "Business %"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths for summary
    for i in range(1, 6):
        ws_summary.column_dimensions[get_column_letter(i)].width = 15
    
    # Add months and formulas
    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    
    for row_idx, month in enumerate(months, 2):
        ws_summary.cell(row=row_idx, column=1).value = month
        
        # These would need to be adjusted based on actual data ranges
        # For now, adding placeholder formulas
        ws_summary.cell(row=row_idx, column=2).value = f'=SUMIFS(\'Mileage Log\'.G:G,\'Mileage Log\'.A:A,">="&DATE(2025,{row_idx-1},1),\'Mileage Log\'.A:A,"<"&DATE(2025,{row_idx},1))'
        ws_summary.cell(row=row_idx, column=3).value = f'=SUMIFS(\'Mileage Log\'.I:I,\'Mileage Log\'.A:A,">="&DATE(2025,{row_idx-1},1),\'Mileage Log\'.A:A,"<"&DATE(2025,{row_idx},1))'
        ws_summary.cell(row=row_idx, column=4).value = f'=SUMIFS(\'Mileage Log\'.J:J,\'Mileage Log\'.A:A,">="&DATE(2025,{row_idx-1},1),\'Mileage Log\'.A:A,"<"&DATE(2025,{row_idx},1))'
        ws_summary.cell(row=row_idx, column=5).value = f'=IF(B{row_idx}>0,C{row_idx}/B{row_idx}*100,0)'
        
        # Add borders
        for col in range(1, 6):
            ws_summary.cell(row=row_idx, column=col).border = border
    
    # Add totals row
    totals_row = len(months) + 2
    ws_summary.cell(row=totals_row, column=1).value = "TOTALS"
    ws_summary.cell(row=totals_row, column=1).font = Font(bold=True)
    
    for col in range(2, 6):
        if col == 5:  # Business percentage
            ws_summary.cell(row=totals_row, column=col).value = f'=IF(B{totals_row}>0,C{totals_row}/B{totals_row}*100,0)'
        else:
            ws_summary.cell(row=totals_row, column=col).value = f'=SUM(B2:B{totals_row-1})' if col == 2 else f'=SUM({get_column_letter(col)}2:{get_column_letter(col)}{totals_row-1})'
        ws_summary.cell(row=totals_row, column=col).font = Font(bold=True)
        ws_summary.cell(row=totals_row, column=col).border = border
    
    # Instructions Sheet
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ["TAXFIX MILEAGE LOG 2025 - INSTRUCTIONS", "", ""],
        ["", "", ""],
        ["How to Use This Template:", "", ""],
        ["", "", ""],
        ["1. DAILY ENTRIES:", "", ""],
        ["   • Enter date in MM/DD/YYYY format", "", ""],
        ["   • Record exact start and end locations", "", ""],
        ["   • Describe the business purpose clearly", "", ""],
        ["   • Enter odometer readings accurately", "", ""],
        ["   • Set business percentage (0-100)", "", ""],
        ["", "", ""],
        ["2. AUTOMATIC CALCULATIONS:", "", ""],
        ["   • Total Miles = End Odometer - Start Odometer", "", ""],
        ["   • Business Miles = Total Miles × Business %", "", ""],
        ["   • Personal Miles = Total Miles - Business Miles", "", ""],
        ["", "", ""],
        ["3. IRS REQUIREMENTS:", "", ""],
        ["   • Record trips on the day they occur", "", ""],
        ["   • Keep receipts for gas, maintenance, insurance", "", ""],
        ["   • Document business purpose for each trip", "", ""],
        ["   • Maintain accurate odometer readings", "", ""],
        ["", "", ""],
        ["4. TAX DEDUCTIONS:", "", ""],
        ["   • Standard Rate 2025: $0.655 per business mile", "", ""],
        ["   • Multiply total business miles × $0.655", "", ""],
        ["   • OR use actual expense method with receipts", "", ""],
        ["", "", ""],
        ["5. MONTHLY SUMMARY:", "", ""],
        ["   • Check the Monthly Summary tab", "", ""],
        ["   • Review totals for accuracy", "", ""],
        ["   • Use for quarterly tax payments", "", ""],
        ["", "", ""],
        ["⚠️ IMPORTANT REMINDERS:", "", ""],
        ["• Keep this log in your vehicle", "", ""],
        ["• Record trips immediately", "", ""],
        ["• Back up your data regularly", "", ""],
        ["• Consult tax professional for complex situations", "", ""],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 1:  # Title
                cell.font = Font(bold=True, size=16, color="4472C4")
            elif ":" in str(value) and not value.startswith(" "):  # Section headers
                cell.font = Font(bold=True, color="4472C4")
            elif value.startswith("•") or value.startswith("⚠️"):  # Important points
                cell.font = Font(bold=True)
    
    # Set column widths for instructions
    ws_instructions.column_dimensions['A'].width = 50
    ws_instructions.column_dimensions['B'].width = 20
    ws_instructions.column_dimensions['C'].width = 20
    
    # Save the workbook
    wb.save("C:/Users/Ben/taxfix-updated-site/TaxFix_Mileage_Log_2025.xlsx")
    print("Mileage Log template created successfully!")

if __name__ == "__main__":
    create_mileage_log()