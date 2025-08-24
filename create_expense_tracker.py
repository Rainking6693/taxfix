from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_expense_tracker():
    wb = Workbook()
    
    # Create main tracking sheet
    ws_main = wb.active
    ws_main.title = "Expense Tracker"
    
    # Headers
    headers = [
        "Date", "Vendor/Description", "Category", "Amount", "Payment Method", 
        "Receipt #", "Business %", "Business Amount", "Personal Amount", "Notes"
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    column_widths = [12, 30, 20, 12, 15, 12, 12, 15, 15, 25]
    for i, width in enumerate(column_widths, 1):
        ws_main.column_dimensions[get_column_letter(i)].width = width
    
    # Add sample data
    sample_data = [
        ["1/1/2025", "Office Depot", "Office Supplies", 89.99, "Credit Card", "RCT001", 100, "", "", "Printer paper, pens"],
        ["1/2/2025", "Starbucks", "Meals & Entertainment", 15.50, "Cash", "", 50, "", "", "Client meeting"],
        ["1/3/2025", "Gas Station", "Vehicle Expenses", 45.00, "Debit Card", "", 80, "", "", "Business travel"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_main.cell(row=row_idx, column=col_idx).value = value
            ws_main.cell(row=row_idx, column=col_idx).border = border
    
    # Add formulas for calculations (rows 2-102 for extensive tracking)
    for row in range(2, 103):
        # Business Amount formula
        ws_main[f'H{row}'] = f'=IF(AND(D{row}<>"",G{row}<>""),D{row}*G{row}/100,"")'
        
        # Personal Amount formula
        ws_main[f'I{row}'] = f'=IF(AND(D{row}<>"",H{row}<>""),D{row}-H{row},"")'
        
        # Add borders
        for col in range(1, 11):
            ws_main.cell(row=row, column=col).border = border
    
    # Create Categories sheet with all 30+ categories
    ws_categories = wb.create_sheet("Categories")
    
    categories = [
        ["BUSINESS EXPENSE CATEGORIES", "DESCRIPTION", "DEDUCTIBLE"],
        ["", "", ""],
        ["Office Supplies", "Paper, pens, computer supplies", "100%"],
        ["Equipment", "Computers, printers, furniture", "100%"],
        ["Software", "Business software subscriptions", "100%"],
        ["Internet & Phone", "Business internet, phone service", "Business %"],
        ["Vehicle Expenses", "Gas, maintenance, insurance", "Business %"],
        ["Meals & Entertainment", "Client meals, business dinners", "50%"],
        ["Travel", "Flights, hotels, rental cars", "100%"],
        ["Professional Services", "Legal, accounting, consulting", "100%"],
        ["Insurance", "Business liability, equipment", "100%"],
        ["Rent", "Office rent, co-working space", "Business %"],
        ["Utilities", "Electricity, water, gas", "Business %"],
        ["Marketing & Advertising", "Website, business cards, ads", "100%"],
        ["Professional Development", "Training, courses, conferences", "100%"],
        ["Subscriptions", "Business magazines, online tools", "100%"],
        ["Bank Fees", "Business banking fees", "100%"],
        ["Licenses & Permits", "Business licenses, permits", "100%"],
        ["Repairs & Maintenance", "Equipment repairs", "100%"],
        ["Shipping", "Postage, shipping supplies", "100%"],
        ["Contract Labor", "Freelancers, contractors", "100%"],
        ["Employee Benefits", "Health insurance, retirement", "100%"],
        ["Taxes", "Business taxes, property tax", "100%"],
        ["Interest", "Business loan interest", "100%"],
        ["Donations", "Charitable contributions", "100%"],
        ["Home Office", "Portion of home expenses", "Business %"],
        ["Tools", "Business tools, equipment", "100%"],
        ["Materials", "Raw materials, supplies", "100%"],
        ["Parking & Tolls", "Business-related parking", "100%"],
        ["Medical", "Business health expenses", "100%"],
        ["Other", "Miscellaneous business expenses", "Varies"],
    ]
    
    # Add category data
    for row_idx, row_data in enumerate(categories, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_categories.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 1:  # Header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            elif row_idx > 2:  # Data rows
                cell.border = border
            cell.alignment = Alignment(horizontal='left')
    
    # Set column widths for categories
    ws_categories.column_dimensions['A'].width = 25
    ws_categories.column_dimensions['B'].width = 35
    ws_categories.column_dimensions['C'].width = 15
    
    # Create Quarterly Summary sheet
    ws_quarterly = wb.create_sheet("Quarterly Summary")
    
    # Quarterly headers
    quarterly_headers = ["Category", "Q1 Total", "Q2 Total", "Q3 Total", "Q4 Total", "Annual Total"]
    for col, header in enumerate(quarterly_headers, 1):
        cell = ws_quarterly.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    for i in range(1, 7):
        ws_quarterly.column_dimensions[get_column_letter(i)].width = 15
    
    # Add category summaries (simplified for template)
    summary_categories = [
        "Office Supplies", "Equipment", "Software", "Vehicle Expenses", 
        "Meals & Entertainment", "Travel", "Professional Services", "Marketing",
        "Home Office", "Other"
    ]
    
    for row_idx, category in enumerate(summary_categories, 2):
        ws_quarterly.cell(row=row_idx, column=1).value = category
        
        # Add formulas for quarterly totals
        for quarter in range(1, 5):
            # These formulas would need to be adjusted based on actual date ranges
            ws_quarterly.cell(row=row_idx, column=quarter+1).value = f'=SUMIFS(\'Expense Tracker\'.H:H,\'Expense Tracker\'.C:C,A{row_idx})'
        
        # Annual total
        ws_quarterly.cell(row=row_idx, column=6).value = f'=SUM(B{row_idx}:E{row_idx})'
        
        # Add borders
        for col in range(1, 7):
            ws_quarterly.cell(row=row_idx, column=col).border = border
    
    # Monthly Summary Sheet
    ws_monthly = wb.create_sheet("Monthly Summary")
    
    monthly_headers = ["Month", "Total Expenses", "Business Expenses", "Personal Expenses", "Receipt Count"]
    for col, header in enumerate(monthly_headers, 1):
        cell = ws_monthly.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    for i in range(1, 6):
        ws_monthly.column_dimensions[get_column_letter(i)].width = 18
    
    # Add months
    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    
    for row_idx, month in enumerate(months, 2):
        ws_monthly.cell(row=row_idx, column=1).value = month
        
        # Add formulas (simplified)
        ws_monthly.cell(row=row_idx, column=2).value = f'=SUMIFS(\'Expense Tracker\'.D:D,\'Expense Tracker\'.A:A,">="&DATE(2025,{row_idx-1},1),\'Expense Tracker\'.A:A,"<"&DATE(2025,{row_idx},1))'
        ws_monthly.cell(row=row_idx, column=3).value = f'=SUMIFS(\'Expense Tracker\'.H:H,\'Expense Tracker\'.A:A,">="&DATE(2025,{row_idx-1},1),\'Expense Tracker\'.A:A,"<"&DATE(2025,{row_idx},1))'
        ws_monthly.cell(row=row_idx, column=4).value = f'=SUMIFS(\'Expense Tracker\'.I:I,\'Expense Tracker\'.A:A,">="&DATE(2025,{row_idx-1},1),\'Expense Tracker\'.A:A,"<"&DATE(2025,{row_idx},1))'
        ws_monthly.cell(row=row_idx, column=5).value = f'=COUNTIFS(\'Expense Tracker\'.A:A,">="&DATE(2025,{row_idx-1},1),\'Expense Tracker\'.A:A,"<"&DATE(2025,{row_idx},1),\'Expense Tracker\'.D:D,">0")'
        
        # Add borders
        for col in range(1, 6):
            ws_monthly.cell(row=row_idx, column=col).border = border
    
    # Instructions Sheet
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ["TAXFIX BUSINESS EXPENSE TRACKER 2025", "", ""],
        ["", "", ""],
        ["How to Use This Template:", "", ""],
        ["", "", ""],
        ["1. RECORDING EXPENSES:", "", ""],
        ["   • Enter date in MM/DD/YYYY format", "", ""],
        ["   • Record vendor name and description", "", ""],
        ["   • Select category from Categories sheet", "", ""],
        ["   • Enter total amount paid", "", ""],
        ["   • Note payment method (cash, card, check)", "", ""],
        ["   • Add receipt number if available", "", ""],
        ["   • Set business percentage (0-100%)", "", ""],
        ["", "", ""],
        ["2. DEDUCTION OPTIMIZATION:", "", ""],
        ["   • Review Categories sheet for guidance", "", ""],
        ["   • Meals = 50% deductible for business", "", ""],
        ["   • Home office = business use percentage", "", ""],
        ["   • Vehicle = business miles percentage", "", ""],
        ["   • 100% business expenses fully deductible", "", ""],
        ["", "", ""],
        ["3. RECEIPT MANAGEMENT:", "", ""],
        ["   • Take photo of all receipts immediately", "", ""],
        ["   • Store receipts for 7 years (IRS requirement)", "", ""],
        ["   • Use receipt number to link to digital copies", "", ""],
        ["   • Keep backup copies in cloud storage", "", ""],
        ["", "", ""],
        ["4. QUARTERLY REVIEWS:", "", ""],
        ["   • Check Quarterly Summary tab", "", ""],
        ["   • Review expense categories for accuracy", "", ""],
        ["   • Calculate quarterly tax payments", "", ""],
        ["   • Plan for upcoming expenses", "", ""],
        ["", "", ""],
        ["5. TAX PREPARATION:", "", ""],
        ["   • Use Monthly/Quarterly summaries", "", ""],
        ["   • Organize receipts by category", "", ""],
        ["   • Provide totals to tax preparer", "", ""],
        ["   • Keep detailed records for audit protection", "", ""],
        ["", "", ""],
        ["⚠️ IMPORTANT REMINDERS:", "", ""],
        ["• Record expenses immediately", "", ""],
        ["• Keep business and personal separate", "", ""],
        ["• Save receipts for everything", "", ""],
        ["• Consult tax professional for guidance", "", ""],
        ["• Back up your data regularly", "", ""],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 1:  # Title
                cell.font = Font(bold=True, size=16, color="70AD47")
            elif ":" in str(value) and not value.startswith(" "):  # Section headers
                cell.font = Font(bold=True, color="70AD47")
            elif value.startswith("•") or value.startswith("⚠️"):  # Important points
                cell.font = Font(bold=True)
    
    # Set column widths for instructions
    ws_instructions.column_dimensions['A'].width = 50
    ws_instructions.column_dimensions['B'].width = 20
    ws_instructions.column_dimensions['C'].width = 20
    
    # Save the workbook
    wb.save("C:/Users/Ben/taxfix-updated-site/TaxFix_Business_Expense_Tracker_2025.xlsx")
    print("Business Expense Tracker template created successfully!")

if __name__ == "__main__":
    create_expense_tracker()