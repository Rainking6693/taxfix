from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date

def create_tax_calendar():
    wb = Workbook()
    
    # Create main calendar sheet
    ws_main = wb.active
    ws_main.title = "2025 Tax Calendar"
    
    # Headers
    headers = ["Date", "Deadline Type", "Description", "Priority", "Notes"]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    high_priority_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    medium_priority_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    low_priority_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    column_widths = [12, 20, 50, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws_main.column_dimensions[get_column_letter(i)].width = width
    
    # Tax deadline data for 2025
    tax_deadlines = [
        # January 2025
        ["1/15/2025", "Quarterly Payment", "Q4 2024 Estimated Tax Payment Due", "HIGH", "Use Form 1040ES"],
        ["1/31/2025", "Annual Deadline", "1099s & W-2s Must be Sent to Recipients", "HIGH", "Businesses must send to workers"],
        ["1/31/2025", "Annual Deadline", "Form 1096 & 1099s Due to IRS (Paper Filing)", "MEDIUM", "Electronic filing deadline is different"],
        
        # February 2025
        ["2/28/2025", "Annual Deadline", "Form 1096 & 1099s Due to IRS (Electronic Filing)", "HIGH", "Electronic filing preferred"],
        
        # March 2025
        ["3/17/2025", "Annual Deadline", "S-Corp Tax Return Due (Form 1120S)", "HIGH", "Can request extension to Sept 15"],
        ["3/17/2025", "Annual Deadline", "Partnership Tax Return Due (Form 1065)", "HIGH", "Can request extension to Sept 15"],
        
        # April 2025
        ["4/15/2025", "Annual Deadline", "Individual Tax Return Due (Form 1040)", "HIGH", "Most important deadline!"],
        ["4/15/2025", "Quarterly Payment", "Q1 2025 Estimated Tax Payment Due", "HIGH", "Use Form 1040ES"],
        ["4/15/2025", "Annual Deadline", "C-Corp Tax Return Due (Form 1120)", "HIGH", "Can request extension to Oct 15"],
        ["4/15/2025", "Annual Deadline", "Request Extension for Individual Returns", "MEDIUM", "Form 4868 - extends to Oct 15"],
        
        # May 2025
        ["5/15/2025", "Annual Deadline", "Tax-Exempt Organization Returns Due", "LOW", "Form 990 series"],
        
        # June 2025
        ["6/16/2025", "Quarterly Payment", "Q2 2025 Estimated Tax Payment Due", "HIGH", "Use Form 1040ES"],
        
        # September 2025
        ["9/15/2025", "Annual Deadline", "Extended S-Corp & Partnership Returns Due", "MEDIUM", "If extension filed in March"],
        ["9/15/2025", "Quarterly Payment", "Q3 2025 Estimated Tax Payment Due", "HIGH", "Use Form 1040ES"],
        
        # October 2025
        ["10/15/2025", "Annual Deadline", "Extended Individual Tax Returns Due", "HIGH", "If extension filed in April"],
        ["10/15/2025", "Annual Deadline", "Extended C-Corp Tax Returns Due", "MEDIUM", "If extension filed in April"],
        
        # December 2025
        ["12/31/2025", "Annual Deadline", "Retirement Plan Contributions", "MEDIUM", "401k, IRA contributions"],
        ["12/31/2025", "Annual Deadline", "HSA Contributions for 2025", "LOW", "Health Savings Account"],
        
        # Additional Important Dates
        ["1/15/2026", "Quarterly Payment", "Q4 2025 Estimated Tax Payment Due", "HIGH", "Following year - mark calendar now"],
    ]
    
    # Add deadline data
    for row_idx, row_data in enumerate(tax_deadlines, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_main.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            
            # Apply priority-based formatting
            priority = row_data[3]
            if priority == "HIGH":
                cell.fill = high_priority_fill
                if col_idx == 4:  # Priority column
                    cell.font = Font(bold=True, color="C62828")
            elif priority == "MEDIUM":
                cell.fill = medium_priority_fill
                if col_idx == 4:  # Priority column
                    cell.font = Font(bold=True, color="F57C00")
            else:  # LOW
                cell.fill = low_priority_fill
                if col_idx == 4:  # Priority column
                    cell.font = Font(bold=True, color="2E7D32")
    
    # Create Monthly View sheet
    ws_monthly = wb.create_sheet("Monthly View")
    
    # Monthly view headers
    monthly_headers = ["Month", "Key Deadlines", "Estimated Tax Due", "Business Returns", "Individual Returns"]
    for col, header in enumerate(monthly_headers, 1):
        cell = ws_monthly.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    for i in range(1, 6):
        ws_monthly.column_dimensions[get_column_letter(i)].width = 20
    
    # Monthly breakdown
    monthly_data = [
        ["JANUARY", "Q4 Estimated Tax (15th), 1099s to Recipients (31st)", "✓", "1099 Filing", "N/A"],
        ["FEBRUARY", "1099s to IRS (28th Electronic)", "N/A", "Info Returns", "N/A"],
        ["MARCH", "S-Corp & Partnership Returns (17th)", "N/A", "✓", "N/A"],
        ["APRIL", "Individual Returns (15th), Q1 Estimated (15th)", "✓", "C-Corp Returns", "✓"],
        ["MAY", "Nonprofit Returns (15th)", "N/A", "Tax-Exempt", "N/A"],
        ["JUNE", "Q2 Estimated Tax (16th)", "✓", "N/A", "N/A"],
        ["JULY", "N/A", "N/A", "N/A", "N/A"],
        ["AUGUST", "N/A", "N/A", "N/A", "N/A"],
        ["SEPTEMBER", "Q3 Estimated Tax (15th), Extended Business Returns", "✓", "✓", "N/A"],
        ["OCTOBER", "Extended Individual Returns (15th)", "N/A", "Extended C-Corp", "✓"],
        ["NOVEMBER", "N/A", "N/A", "N/A", "N/A"],
        ["DECEMBER", "Retirement Contributions, HSA Contributions", "N/A", "N/A", "Planning"],
    ]
    
    for row_idx, row_data in enumerate(monthly_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_monthly.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            if col_idx == 1:  # Month column
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Create State Deadlines sheet
    ws_state = wb.create_sheet("State Deadlines")
    
    state_info = [
        ["STATE TAX DEADLINE VARIATIONS", "", ""],
        ["", "", ""],
        ["Most states follow federal deadlines, but some variations exist:", "", ""],
        ["", "", ""],
        ["DIFFERENT STATE DEADLINES:", "", ""],
        ["Delaware", "April 30th (Individual)", ""],
        ["Hawaii", "April 20th (Individual)", ""],
        ["Iowa", "April 30th (Individual)", ""],
        ["Louisiana", "May 15th (Individual)", ""],
        ["Maine", "April 19th (if April 15th falls on weekend)", ""],
        ["", "", ""],
        ["QUARTERLY ESTIMATED TAX PAYMENTS:", "", ""],
        ["Most states follow federal schedule:", "", ""],
        ["• Q1: April 15th", "", ""],
        ["• Q2: June 15th", "", ""],
        ["• Q3: September 15th", "", ""],
        ["• Q4: January 15th (following year)", "", ""],
        ["", "", ""],
        ["IMPORTANT NOTES:", "", ""],
        ["• Check your specific state's tax website", "", ""],
        ["• Some states have no income tax", "", ""],
        ["• Business tax deadlines may vary", "", ""],
        ["• Extension deadlines usually match federal", "", ""],
        ["• Local taxes may have different deadlines", "", ""],
    ]
    
    for row_idx, row_data in enumerate(state_info, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_state.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 1:  # Title
                cell.font = Font(bold=True, size=16, color="E74C3C")
            elif ":" in str(value) and not value.startswith(" ") and not value.startswith("•"):
                cell.font = Font(bold=True, color="E74C3C")
            elif value.startswith("•"):
                cell.font = Font(italic=True)
    
    # Set column widths
    ws_state.column_dimensions['A'].width = 40
    ws_state.column_dimensions['B'].width = 30
    ws_state.column_dimensions['C'].width = 20
    
    # Create Instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ["TAXFIX 2025 TAX DEADLINE CALENDAR", "", ""],
        ["", "", ""],
        ["How to Use This Calendar:", "", ""],
        ["", "", ""],
        ["1. PRIORITY SYSTEM:", "", ""],
        ["   • HIGH = Critical deadlines, penalties if missed", "", ""],
        ["   • MEDIUM = Important but extensions available", "", ""],
        ["   • LOW = Helpful dates, less critical", "", ""],
        ["", "", ""],
        ["2. QUARTERLY ESTIMATED TAXES:", "", ""],
        ["   • Due: Jan 15, Apr 15, Jun 15, Sep 15", "", ""],
        ["   • Use Form 1040ES", "", ""],
        ["   • Pay 25% of estimated annual tax each quarter", "", ""],
        ["   • Avoid underpayment penalties", "", ""],
        ["", "", ""],
        ["3. BUSINESS TAX RETURNS:", "", ""],
        ["   • S-Corp & Partnerships: March 15th", "", ""],
        ["   • C-Corporations: April 15th", "", ""],
        ["   • Extensions available but taxes still due", "", ""],
        ["", "", ""],
        ["4. INDIVIDUAL TAX RETURNS:", "", ""],
        ["   • Deadline: April 15th", "", ""],
        ["   • Extension to October 15th with Form 4868", "", ""],
        ["   • Extensions don't extend payment deadline", "", ""],
        ["", "", ""],
        ["5. INFORMATION RETURNS:", "", ""],
        ["   • 1099s to recipients: January 31st", "", ""],
        ["   • 1099s to IRS: February 28th (electronic)", "", ""],
        ["   • W-2s to employees: January 31st", "", ""],
        ["", "", ""],
        ["6. CALENDAR INTEGRATION:", "", ""],
        ["   • Export dates to Google Calendar", "", ""],
        ["   • Set reminders 2 weeks before deadlines", "", ""],
        ["   • Add preparation time before due dates", "", ""],
        ["", "", ""],
        ["⚠️ PENALTY AVOIDANCE TIPS:", "", ""],
        ["• File extensions even if you can't pay", "", ""],
        ["• Pay estimated taxes to avoid penalties", "", ""],
        ["• Keep detailed records year-round", "", ""],
        ["• Consider professional help for complex situations", "", ""],
        ["• Review state-specific deadlines", "", ""],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instructions.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 1:  # Title
                cell.font = Font(bold=True, size=16, color="E74C3C")
            elif ":" in str(value) and not value.startswith(" "):
                cell.font = Font(bold=True, color="E74C3C")
            elif value.startswith("•") or value.startswith("⚠️"):
                cell.font = Font(bold=True)
    
    # Set column widths
    ws_instructions.column_dimensions['A'].width = 50
    ws_instructions.column_dimensions['B'].width = 20
    ws_instructions.column_dimensions['C'].width = 20
    
    # Save the workbook
    wb.save("C:/Users/Ben/taxfix-updated-site/TaxFix_2025_Tax_Deadline_Calendar.xlsx")
    print("2025 Tax Deadline Calendar created successfully!")

if __name__ == "__main__":
    create_tax_calendar()